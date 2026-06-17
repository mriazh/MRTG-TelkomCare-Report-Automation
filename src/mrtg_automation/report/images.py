import io
import logging
from pathlib import Path
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Constants for Excel pixel conversions (approximate)
PX_CONV_WIDTH = 7.4
PX_CONV_HEIGHT = 1.333
EXCEL_DEFAULT_WIDTH = 8.43
EXCEL_DEFAULT_HEIGHT = 15.0

def get_area_size_pixels(sheet, start_row: int, start_col: int, end_row: int, end_col: int):
    """
    Calculate the pixel size of a specified Excel cell range.
    """
    total_width = 0
    for col in range(start_col, end_col + 1):
        col_letter = get_column_letter(col)
        col_width = sheet.column_dimensions[col_letter].width
        if col_width is None:
            col_width = EXCEL_DEFAULT_WIDTH
        total_width += col_width * PX_CONV_WIDTH

    total_height = 0
    for row in range(start_row, end_row + 1):
        row_height = sheet.row_dimensions[row].height
        if row_height is None:
            row_height = EXCEL_DEFAULT_HEIGHT
        total_height += row_height * PX_CONV_HEIGHT

    return total_width, total_height

def resize_image_stretch(image_path: Path, target_width: float, target_height: float) -> io.BytesIO:
    """
    Resize an image to stretch into the target dimensions.
    Returns the image data as a BytesIO object.
    """
    with PILImage.open(image_path) as img:
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGB')
        
        # LANCZOS is high quality downsampling
        img_resized = img.resize((int(target_width), int(target_height)), PILImage.Resampling.LANCZOS)
        
        output = io.BytesIO()
        img_resized.save(output, format='PNG')
        output.seek(0)
        return output

def insert_image_to_area(sheet, image_path: Path, start_row: int, start_col: int, end_row: int, end_col: int, scale: float = 0.98) -> bool:
    """
    Resize and insert an image into the specified Excel cell area.
    """
    try:
        width_px, height_px = get_area_size_pixels(sheet, start_row, start_col, end_row, end_col)
        width_px = width_px * scale
        height_px = height_px * scale
        
        img_bytes = resize_image_stretch(image_path, width_px, height_px)
        img = XLImage(img_bytes)
        
        # Anchor top-left to start_col and start_row
        img.anchor = f"{get_column_letter(start_col)}{start_row}"
        sheet.add_image(img)
        return True
    except Exception as e:
        logger.warning(f"Failed to insert image {image_path.name}: {e}")
        return False
