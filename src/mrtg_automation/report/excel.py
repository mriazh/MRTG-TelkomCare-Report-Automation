import os
import logging
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .mapping import parse_image_only_mapping, parse_target_list, parse_ocr_mapping
from .images import insert_image_to_area
from ..shared.filenames import build_canonical_filename

logger = logging.getLogger(__name__)

class ExcelReportGenerator:
    def __init__(self, config):
        self.config = config

    def _get_tanggal_folders(self, data_dir: Path) -> list:
        """Find and return sorted YYYYMMDD folder names inside the data_dir."""
        if not data_dir.exists():
            return []

        folders = []
        for entry in os.scandir(data_dir):
            if entry.is_dir() and entry.name.isdigit() and len(entry.name) == 8:
                folders.append(entry.name)

        folders.sort()
        return folders

    def generate(self, report_mode: str, data_dir: Path, template_path: Path, output_path: Path, mapping_file: Path, list_file: Path, date_filter: str = None, progress_callback=None, cancel_event=None, resume_state=None, resume_mode: bool = False, phase: str = None):
        """
        Main orchestration logic for report generation.
        """
        summary = {
            "success": False,
            "dates_processed": 0,
            "targets": 0,
            "expected": 0,
            "image_inserted": 0,
            "missing_screenshots": 0,
            "missing_mappings": 0,
            "failed_inserts": 0,
            "ocr_ok": 0,
            "ocr_partial": 0,
            "ocr_fail": 0,
            "review_list": [],
            "output_file": str(output_path)
        }

        if report_mode not in ('IMAGE_ONLY', 'OCR_IMAGE'):
            logger.error(f"Mode {report_mode} is not yet implemented.")
            return summary

        if not template_path.exists():
            logger.error(f"Template Excel not found: {template_path}")
            return summary

        if not mapping_file.exists():
            logger.error(f"Mapping file not found: {mapping_file}")
            return summary

        if not list_file.exists():
            logger.error(f"Target list file not found: {list_file}")
            return summary

        if not data_dir.exists():
            logger.error(f"Data directory not found: {data_dir}")
            return summary

        # Load mapping based on mode
        if report_mode == 'IMAGE_ONLY':
            mapping = parse_image_only_mapping(mapping_file)
        elif report_mode == 'OCR_IMAGE':
            mapping = parse_ocr_mapping(mapping_file)

        if not mapping:
            logger.error("Mapping is empty. Aborting.")
            return summary

        target_filter = "image" if report_mode == "IMAGE_ONLY" else "ocr"
        items = parse_target_list(list_file, enabled_for=target_filter)
        if not items:
            logger.error("Target list is empty. Aborting.")
            return summary

        summary["targets"] = len(items)
        logger.info(f"Loaded {len(mapping)} mappings and {len(items)} target items.")

        tanggal_list = [d.name for d in data_dir.iterdir() if d.is_dir() and d.name.isdigit()]
        if date_filter:
            if isinstance(date_filter, str):
                tanggal_list = [t for t in tanggal_list if t == date_filter]
            else:
                allowed = set(date_filter)
                tanggal_list = [t for t in tanggal_list if t in allowed]

        tanggal_list.sort()
        if not tanggal_list:
            logger.error(f"No valid YYYYMMDD folders found in {data_dir} matching filter.")
            return summary

        summary["dates_processed"] = len(tanggal_list)
        summary["expected"] = len(items) * len(tanggal_list)

        phase = phase or ("report_image" if report_mode == "IMAGE_ONLY" else "report_ocr")

        if resume_mode and output_path.exists():
            logger.info(f"Loading partial workbook: {output_path.name}")
            print(f"[RESUME] Loading partial workbook: {output_path}")
            wb = load_workbook(output_path)
        else:
            logger.info(f"Loading template: {template_path.name}")
            wb = load_workbook(template_path)

        # Lazy load OCRExtractor
        OCRExtractor = None
        if report_mode == 'OCR_IMAGE':
            try:
                from .ocr import OCRExtractor
                # Warm up the engine to catch failures early
                OCRExtractor._get_engine()
            except ImportError as e:
                logger.error(str(e))
                print(f"\n[FAIL] {e}")
                return summary
            except Exception as e:
                logger.error(str(e))
                print(f"\n[FAIL] {e}")
                return summary

        total_items = len(tanggal_list) * len(items)
        current_index = 0

        from mrtg_automation.shared.resume_state import get_completed_item_keys, mark_item_completed, save_resume_state, make_item_key, count_completed_items_for_phase
        if resume_state is not None:
            resume_state["status"] = "running"
            resume_state["current_phase"] = phase
            resume_state["phase_total_items"] = total_items
            resume_state["phase_completed_items_count"] = count_completed_items_for_phase(resume_state, phase)
            if not resume_state.get("total_items"):
                resume_state["total_items"] = total_items
            save_resume_state(resume_state)

        completed_keys = get_completed_item_keys(resume_state) if resume_state and resume_mode else set()

        def find_completed_item(key):
            if not resume_state:
                return None
            for item in resume_state.get("completed_items", []):
                if item.get("key") == key:
                    return item
            return None

        # Loop through each date folder
        for tanggal_str in tanggal_list:
            hari = int(tanggal_str[6:8])
            sheet_name = f"{hari:02d}"

            # Find the target sheet (01, 02.. or 1, 2..)
            if sheet_name not in wb.sheetnames:
                sheet_name = str(hari)
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"Sheet {hari:02d} not found in template. Skipping date {tanggal_str}.")
                    continue

            sheet = wb[sheet_name]
            logger.info(f"Processing date: {tanggal_str} on Sheet: {sheet_name}")

            # Parse the string into a datetime object to build the canonical filename
            try:
                date_obj = datetime.strptime(tanggal_str, "%Y%m%d")
            except ValueError:
                logger.warning(f"Invalid date format for folder {tanggal_str}. Skipping.")
                continue

            for nomor, tipe, target_id in items:
                key = make_item_key(phase, report_mode, tanggal_str, target_id)

                if resume_state is not None:
                    resume_state["next_item"] = {
                        "phase": phase,
                        "mode": report_mode,
                        "date": tanggal_str,
                        "target": target_id,
                        "key": key
                    }
                    save_resume_state(resume_state)

                if cancel_event is not None and cancel_event.is_set():
                    print("[STOP] Report stop requested. Saving partial workbook before stopping.")
                    logger.warning("[STOP] Report stop requested. Stopping before next item.")
                    output_path.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(output_path)
                    summary["cancelled"] = True
                    summary["success"] = False
                    if resume_state is not None:
                        resume_state["status"] = "stopped"
                        save_resume_state(resume_state)
                    return summary

                if resume_mode and key in completed_keys:
                    print(f"[SKIP] report {current_index + 1}/{total_items} mode={report_mode} date={tanggal_str} target={target_id} already completed")
                    current_index += 1
                    continue

                current_index += 1
                prog_msg = f"[PROGRESS] report {current_index}/{total_items} mode={report_mode} date={tanggal_str} target={target_id} starting"
                print(prog_msg)
                logger.info(prog_msg)

                item_status = "OK"
                item_error = None
                item_suffix = ""
                ocr_status_for_state = None

                if target_id not in mapping:
                    msg = f"[N/A] report {current_index}/{total_items} mode={report_mode} date={tanggal_str} target={target_id} reason=not_in_mapping"
                    print(msg)
                    logger.warning(msg)
                    summary["missing_mappings"] += 1
                    if resume_state is not None:
                        item = {"phase": phase, "mode": report_mode, "date": tanggal_str, "target": target_id, "status": "missing_mapping", "error": None, "key": key}
                        mark_item_completed(resume_state, item)
                        resume_state["phase_completed_items_count"] = count_completed_items_for_phase(resume_state, phase)
                        save_resume_state(resume_state)
                        completed_keys.add(key)
                    continue

                if report_mode == 'IMAGE_ONLY':
                    (start_row, start_col), (end_row, end_col) = mapping[target_id]
                else:
                    if 'Image' in mapping[target_id]:
                        (start_row, start_col), (end_row, end_col) = mapping[target_id]['Image']
                    else:
                        start_row, start_col, end_row, end_col = None, None, None, None

                # Strict requirement: only use canonical format
                filename = build_canonical_filename(target_id, date_obj)
                path_gambar = data_dir / tanggal_str / filename

                if not path_gambar.exists():
                    msg = f"[FAIL] report {current_index}/{total_items} mode={report_mode} date={tanggal_str} target={target_id} error=missing_screenshot"
                    print(msg)
                    logger.warning(msg)
                    summary["missing_screenshots"] += 1
                    if resume_state is not None:
                        item = {"phase": phase, "mode": report_mode, "date": tanggal_str, "target": target_id, "status": "missing_screenshot", "error": None, "key": key}
                        mark_item_completed(resume_state, item)
                        resume_state["phase_completed_items_count"] = count_completed_items_for_phase(resume_state, phase)
                        save_resume_state(resume_state)
                        completed_keys.add(key)
                    continue

                # Try OCR if applicable
                if report_mode == 'OCR_IMAGE':
                    print(f"  [{int(nomor):02d}/{len(items)}] {tipe} {target_id} ... ", end="", flush=True)
                    try:
                        ocr_vals = OCRExtractor.extract_mrtg_values(path_gambar)
                    except Exception as e:
                        logger.error(f"OCR Extractor exception: {e}")
                        ocr_vals = None

                    if ocr_vals is None:
                        print("FAIL")
                        summary["ocr_fail"] += 1
                        summary["review_list"].append({
                            "target_id": target_id,
                            "date": tanggal_str,
                            "sheet": sheet_name,
                            "status": "Fail",
                            "na_count": 6
                        })
                        item_status = "FAIL"
                        item_error = "ocr_failed"
                        ocr_status_for_state = "ocr_failed"
                        # Optional: fill all mapped text fields with N/A
                        for field_key, (r, c) in mapping[target_id].items():
                            if field_key != 'Image':
                                sheet.cell(row=r, column=c, value="N/A")
                    else:
                        na_count = sum(1 for v in ocr_vals.values() if v == 'N/A')
                        if na_count == 0:
                            print("OK")
                            summary["ocr_ok"] += 1
                            ocr_status_for_state = "ocr_ok"
                        else:
                            print(f"PARTIAL ({na_count}/6 N/A)")
                            summary["ocr_partial"] += 1
                            ocr_status_for_state = "ocr_partial"
                            summary["review_list"].append({
                                "target_id": target_id,
                                "date": tanggal_str,
                                "sheet": sheet_name,
                                "status": "Partial",
                                "na_count": na_count
                            })
                            item_suffix = f" ocr_status=partial na_count={na_count}"

                        for field_key, (r, c) in mapping[target_id].items():
                            if field_key != 'Image' and field_key in ocr_vals:
                                sheet.cell(row=r, column=c, value=ocr_vals[field_key])

                # Insert the image
                insert_img_flag = os.environ.get('INSERT_IMAGES', 'True').lower() == 'true'
                if start_row is not None and insert_img_flag:
                    success = insert_image_to_area(sheet, path_gambar, start_row, start_col, end_row, end_col)
                    if success:
                        logger.info(f"  Inserted {path_gambar.name} into {get_column_letter(start_col)}{start_row}")
                        summary["image_inserted"] += 1
                    else:
                        summary["failed_inserts"] += 1
                        item_status = "FAIL"
                        if not item_error:
                            item_error = "image_insert_failed"
                        else:
                            item_error += ",image_insert_failed"

                if item_status != "OK":
                    status_for_state = item_error if item_error else "error"
                    msg = f"[FAIL] report {current_index}/{total_items} mode={report_mode} date={tanggal_str} target={target_id} error={item_error}{item_suffix}"
                    print(msg)
                    logger.error(msg)
                elif report_mode == "OCR_IMAGE" and ocr_status_for_state:
                    status_for_state = ocr_status_for_state
                    msg = f"[OK] report {current_index}/{total_items} mode={report_mode} date={tanggal_str} target={target_id}{item_suffix}"
                    print(msg)
                    logger.info(msg)
                else:
                    status_for_state = "ok"
                    msg = f"[OK] report {current_index}/{total_items} mode={report_mode} date={tanggal_str} target={target_id}{item_suffix}"
                    print(msg)
                    logger.info(msg)

                if resume_state is not None:
                    item = {
                        "phase": phase,
                        "mode": report_mode,
                        "date": tanggal_str,
                        "target": target_id,
                        "status": status_for_state,
                        "error": item_error,
                        "key": key
                    }
                    mark_item_completed(resume_state, item)
                    resume_state["phase_completed_items_count"] = count_completed_items_for_phase(resume_state, phase)
                    save_resume_state(resume_state)
                    completed_keys.add(key)

        logger.info(f"Saving workbook to {output_path}")
        # Ensure parent directories exist
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info("Report generation complete.")
        summary["success"] = True

        if resume_state is not None:
            resume_state["current_phase"] = phase
            resume_state["next_item"] = None
            resume_state["phase_completed_items_count"] = count_completed_items_for_phase(resume_state, phase)
            save_resume_state(resume_state)

        return summary

    @staticmethod
    def format_summary_table(summary: dict) -> str:
        """Format a human-readable summary table for console output."""
        lines = []
        lines.append("")
        lines.append("=" * 70)
        lines.append(f"  OCR Summary: OK={summary['ocr_ok']} | Partial={summary['ocr_partial']} | Fail={summary['ocr_fail']}")
        lines.append(f"  Image inserted: {summary['image_inserted']} | Missing: {summary['missing_screenshots']}")
        lines.append(f"  Date filter: {summary.get('date_filter', 'N/A')}")
        if summary.get('review_list'):
            lines.append("")
            lines.append(f"  Review list ({len(summary['review_list'])} items):")
            for r in summary['review_list']:
                status_icon = {"Fail": "[FAIL]", "Partial": "[WARN]"}.get(r['status'], "      ")
                lines.append(f"    {status_icon} [{r['status']:7s}] {r['target_id']:25s} @ {r['date']} sheet={r['sheet']} (N/A: {r['na_count']})")
        lines.append("=" * 70)
        return "\n".join(lines)
